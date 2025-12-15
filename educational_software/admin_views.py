from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.utils import timezone

from openpyxl import load_workbook

from .models import Laboratory, Software, SoftwareInstallation


def _is_installed_cell(cell) -> bool:
    """
    Zwraca True tylko dla komórek oznaczonych kolorem (zielony/żółty itp.).
    Kluczowe: odrzucamy białe wypełnienia (Excel często zapisuje białe jako "solid").
    """
    fill = getattr(cell, "fill", None)
    if not fill:
        return False

    # Jeżeli brak wzoru wypełnienia -> traktujemy jako brak instalacji
    if not getattr(fill, "patternType", None):
        return False

    fg = getattr(fill, "fgColor", None)
    if not fg:
        return False

    # Najczęściej interesuje nas RGB. Jeżeli brak RGB (np. theme), nie uznajemy za instalację.
    rgb = getattr(fg, "rgb", None)
    if not rgb:
        return False

    rgb = str(rgb).upper()

    # Odrzucamy "białe" i "puste" warianty, które często robią fałszywe pozytywy.
    WHITE_SET = {
        "FFFFFFFF",  # biały
        "00FFFFFF",  # czasem bez alfa
        "00000000",  # przezroczysty/pusty
        "FF000000",  # czarny (gdyby Excel stylował tabelę)
        "000000",    # czasem krótszy zapis
        "FFFFFF",
    }
    if rgb in WHITE_SET:
        return False

    return True


@staff_member_required(login_url="/admin/login/")
def software_excel_import_view(request):
    """
    IMPORT EXCELA – ZAKŁADKA OPROGRAMOWANIE

    - Czyta TYLKO arkusz „Zmienne programy”
    - Kolumna A = nazwa oprogramowania
    - Kolumny B–S = laboratoria (numery)
    - Zielone / żółte pole = oprogramowanie zainstalowane w danym laboratorium
    - KAŻDY import:
        • usuwa poprzednie instalacje
        • odtwarza stan dokładnie z Excela
    """
    context = {}

    if request.method == "POST" and request.FILES.get("file"):
        excel_file = request.FILES["file"]

        try:
            wb = load_workbook(excel_file, data_only=True)
        except Exception as exc:
            context["error"] = f"Błąd odczytu pliku Excel: {exc}"
            return render(request, "admin/educational_software/import_excel.html", context)

        if "Zmienne programy" not in wb.sheetnames:
            context["error"] = "Brak arkusza „Zmienne programy” w pliku."
            return render(request, "admin/educational_software/import_excel.html", context)

        ws = wb["Zmienne programy"]
        rows = list(ws.iter_rows())

        if len(rows) < 2:
            context["error"] = "Arkusz nie zawiera danych."
            return render(request, "admin/educational_software/import_excel.html", context)

        # 1) Laboratoria z nagłówków (kolumny B..)
        header_row = rows[0]
        laboratories = []
        for cell in header_row[1:]:
            if cell.value:
                lab_number = str(cell.value).strip()
                lab, _ = Laboratory.objects.get_or_create(number=lab_number)
                laboratories.append(lab)
            else:
                laboratories.append(None)

        # 2) Nadpisujemy całość: czyścimy instalacje
        SoftwareInstallation.objects.all().delete()

        software_created = 0
        installations_created = 0

        # 3) Wiersze z oprogramowaniem
        for row in rows[1:]:
            software_name = row[0].value
            if not software_name:
                continue

            software_name = str(software_name).strip()
            software, created = Software.objects.get_or_create(name=software_name)
            if created:
                software_created += 1

            for idx, cell in enumerate(row[1:]):
                lab = laboratories[idx] if idx < len(laboratories) else None
                if not lab:
                    continue

                if _is_installed_cell(cell):
                    SoftwareInstallation.objects.create(
                        software=software,
                        laboratory=lab,
                        status="installed",
                        updated_at=timezone.now(),
                    )
                    installations_created += 1

        context.update(
            {
                "import_done": True,
                "software_created": software_created,
                "installations_created": installations_created,
            }
        )

        return render(request, "admin/educational_software/import_excel.html", context)

    return render(request, "admin/educational_software/import_excel.html", context)
