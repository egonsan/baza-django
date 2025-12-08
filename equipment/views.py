from datetime import date

from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.generic import ListView, DetailView, UpdateView

from openpyxl import load_workbook, Workbook

from .models import Equipment, EquipmentAttachment


# ============================================================
# LISTA SPRZĘTU – MAGAZYN
# ============================================================


@method_decorator(login_required(login_url="/baza/"), name="dispatch")
class EquipmentListView(ListView):
    """
    Lista sprzętu w MAGAZYNIE – /baza/magazyn/

    Logika:
    - pokazuje wyłącznie sprzęt z room_category = "MAGAZYN",
    - proste wyszukiwanie po numerze inwentarzowym, nazwie, użytkowniku, budynku, pokoju.
    """

    model = Equipment
    template_name = "equipment/equipment_list.html"
    context_object_name = "equipments"
    paginate_by = 50

    def get_queryset(self):
        # Tylko sprzęt z kategorii MAGAZYN
        qs = Equipment.objects.filter(room_category="MAGAZYN").order_by(
            "inventory_number"
        )
        q = self.request.GET.get("q", "").strip()

        if q:
            qs = qs.filter(
                Q(inventory_number__icontains=q)
                | Q(equipment_name__icontains=q)
                | Q(hostname__icontains=q)
                | Q(user_full_name__icontains=q)
                | Q(building__icontains=q)
                | Q(room__icontains=q)
            )
        return qs


# ============================================================
# SZCZEGÓŁY SPRZĘTU
# ============================================================


@method_decorator(login_required(login_url="/baza/"), name="dispatch")
class EquipmentDetailView(DetailView):
    """
    Szczegóły sprzętu /baza/magazyn/<pk>/
    Pokazuje też listę załączników.
    """

    model = Equipment
    template_name = "equipment/equipment_detail.html"
    context_object_name = "equipment"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Załączniki powiązane z tym sprzętem
        # USTALAMY KOLEJNOŚĆ: najstarszy jako pierwszy (do zdjęcia),
        # kolejne pliki dopisywane NA KOŃCU listy
        context["attachments"] = EquipmentAttachment.objects.filter(
            equipment=self.object
        ).order_by("uploaded_at")
        return context


# ============================================================
# EDYCJA SPRZĘTU
# ============================================================


@method_decorator(login_required(login_url="/baza/"), name="dispatch")
class EquipmentUpdateView(UpdateView):
    """
    Edycja sprzętu /baza/magazyn/<pk>/edit/
    """

    model = Equipment
    template_name = "equipment/equipment_form.html"
    context_object_name = "equipment"

    # Pola edytowane w formularzu
    fields = [
        "inventory_number",
        "equipment_name",
        "equipment_type",
        "status",
        "hostname",
        "user_full_name",
        "borrowed_to",
        "building",
        "room",
        "ip_address",
        "mac_address",
        "unit_serial_number",
        "monitor_serial_number",
        "os_name",
        "os_version",
        "os_serial_key",
        "office_name",
        "office_version",
        "office_serial_key",
        "supplier",
        "purchase_date",
        "warranty_until",
        "notes",
    ]

    success_url = reverse_lazy("equipment:equipment_list")

    def form_valid(self, form):
        """
        Ustawiamy last_modified_by automatycznie na aktualnie zalogowanego użytkownika.
        """
        obj = form.save(commit=False)
        user = self.request.user if self.request.user.is_authenticated else None
        obj.last_modified_by = user
        obj.save()
        return redirect(self.get_success_url())


# ============================================================
# IMPORT Z EXCELA – WERSJA DLA ADMINA
# ============================================================


@staff_member_required(login_url="/admin/login/")
def admin_equipment_import_view(request):
    """
    Widok importu danych z pliku XLSX wywoływany z panelu admina.

    Adres (zgodnie z urls.py):
    /baza/admin-import/ lub /baza/import/

    Szablon:
    templates/admin/equipment/equipment/import_excel.html

    Założenia Excela:
    - Pierwszy wiersz = nagłówki kolumn
    - Standardowo: nagłówki odpowiadają nazwom pól modelu Equipment
      (inventory_number, equipment_name, itp.)
    - Dodatkowo obsługujemy nagłówek 'DATA_ZAKUP', który mapujemy na
      pole 'purchase_date' (Data zakupu).
    - Rekordy identyfikujemy po 'inventory_number'.
    """

    context = {}

    if request.method == "POST" and request.FILES.get("file"):
        file_obj = request.FILES["file"]

        try:
            wb = load_workbook(file_obj, data_only=True)
        except Exception as exc:
            context["error"] = f"Nie udało się odczytać pliku XLSX: {exc}"
            return render(
                request,
                "admin/equipment/equipment/import_excel.html",
                context,
            )

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            context["error"] = "Plik jest pusty."
            return render(
                request,
                "admin/equipment/equipment/import_excel.html",
                context,
            )

        # Pierwszy wiersz = nagłówki
        header_row = rows[0]
        headers = []
        for cell in header_row:
            if cell is None:
                headers.append("")
            else:
                headers.append(str(cell).strip())

        # Dostępne pola modelu (bez relacji)
        model_fields = {
            f.name
            for f in Equipment._meta.get_fields()
            if f.concrete and not f.many_to_many and not f.auto_created
        }

        created_count = 0
        updated_count = 0
        processed_rows = 0

        # Wiersze danych (od drugiego wiersza)
        for row in rows[1:]:
            # Pomijamy całkowicie puste wiersze
            if all(cell is None for cell in row):
                continue

            processed_rows += 1
            data = {}

            for idx, value in enumerate(row):
                if idx >= len(headers):
                    continue

                original_header = headers[idx]
                if not original_header:
                    continue

                # Mapowanie specjalne: DATA_ZAKUP -> purchase_date
                if original_header.upper() == "DATA_ZAKUP":
                    field_name = "purchase_date"
                else:
                    field_name = original_header

                # Jeżeli po mapowaniu nazwa pola nie istnieje w modelu – pomijamy
                if field_name not in model_fields:
                    continue

                data[field_name] = value

            inv_raw = data.get("inventory_number")
            inventory_number = str(inv_raw).strip() if inv_raw is not None else ""

            if not inventory_number:
                # Bez numeru inwentarzowego nie zapisujemy wiersza
                continue

            # Czyszczenie stringów
            clean_data = {}
            for key, value in data.items():
                if isinstance(value, str):
                    clean_data[key] = value.strip()
                else:
                    clean_data[key] = value

            # Tworzymy lub aktualizujemy rekord po inventory_number
            obj, created = Equipment.objects.update_or_create(
                inventory_number=inventory_number,
                defaults=clean_data,
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        context.update(
            {
                "import_done": True,
                "processed_rows": processed_rows,
                "created_count": created_count,
                "updated_count": updated_count,
            }
        )

        return render(
            request,
            "admin/equipment/equipment/import_excel.html",
            context,
        )

    # GET lub brak pliku – pokazujemy pusty formularz
    return render(
        request,
        "admin/equipment/equipment/import_excel.html",
        context,
    )


# (Opcjonalnie) stara wersja importu dostępna pod /baza/import/
# Jeżeli gdzieś jeszcze używasz tego adresu, możesz zostawić ten widok
# i dodać do urls.py odpowiednią ścieżkę.
@login_required(login_url="/baza/")
def equipment_import_view(request):
    """
    STARSZA wersja importu (nie używana przez admin change_list).
    Zostawiona na wszelki wypadek.
    """
    return admin_equipment_import_view(request)


# ============================================================
# EKSPORT DO EXCELA – WERSJA DLA ADMINA
# ============================================================


@staff_member_required(login_url="/admin/login/")
def admin_equipment_export_view(request):
    """
    Eksport danych do pliku XLSX w formacie zgodnym z używanym Excelem
    (polskie nagłówki: NR_INWENTARZOWY, BUDYNEK, POMIESZCZENIE, itd.).

    Widok przeznaczony do wywoływania z panelu admina / strony importu.
    """

    # Tworzymy nowy skoroszyt
    wb = Workbook()
    ws = wb.active
    ws.title = "Sprzet"

    # Nagłówki w dokładnie takiej formie, jak w pliku importu
    headers = [
        "NR_INWENTARZOWY",
        "BUDYNEK",
        "POMIESZCZENIE",
        "TYP_SPRZETU",
        "NAZWA",
        "NAZWISKO",
        "NR_FABR",
        "NR_SERYJNY_MONITORA",
        "DATA_ZAKUP",
        "UWAGI",
        "KLUCZ_WINDOWS",
        "KLUCZ_OFFICE",
        "MAC_JEDNOSTKI",
        "NAZWA_DOMENOWA",
        "ADRES_IP",
    ]
    ws.append(headers)

    # Pobieramy wszystkie rekordy sprzętu, posortowane po numerze inwentarzowym
    queryset = Equipment.objects.all().order_by("inventory_number")

    for eq in queryset:
        row = [
            eq.inventory_number or "",
            eq.building or "",
            eq.room or "",
            eq.equipment_type or "",
            eq.equipment_name or "",
            eq.user_full_name or "",
            eq.unit_serial_number or "",
            eq.monitor_serial_number or "",
            eq.purchase_date.strftime("%Y-%m-%d") if eq.purchase_date else "",
            eq.notes or "",
            eq.os_serial_key or "",
            eq.office_serial_key or "",
            eq.mac_address or "",
            eq.hostname or "",
            eq.ip_address or "",
        ]
        ws.append(row)

    # Przygotowujemy odpowiedź HTTP z plikiem XLSX
    response = HttpResponse(
        content_type=(
            "application/vnd."
            "openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = 'attachment; filename="karty_sprzetu.xlsx"'

    from io import BytesIO

    output = BytesIO()
    wb.save(output)
    response.write(output.getvalue())

    return response


# ============================================================
# ZAŁĄCZNIKI – DODAWANIE
# ============================================================


@login_required(login_url="/baza/")
def attachment_upload_view(request, pk):
    """
    Upload pojedynczego załącznika do sprzętu.
    Adres: /baza/magazyn/<pk>/upload/
    """

    equipment = get_object_or_404(Equipment, pk=pk)

    if request.method != "POST":
        return HttpResponseForbidden("Niedozwolone żądanie.")

    if "file" not in request.FILES:
        # Brak pliku – wracamy do szczegółów
        return redirect("equipment:equipment_detail", pk=pk)

    EquipmentAttachment.objects.create(
        equipment=equipment,
        file=request.FILES["file"],
        description=request.POST.get("description", "").strip(),
    )

    return redirect("equipment:equipment_detail", pk=equipment.pk)


# ============================================================
# ZAŁĄCZNIKI – USUWANIE
# ============================================================


@login_required(login_url="/baza/")
def attachment_delete_view(request, attachment_id):
    """
    Usuwanie pojedynczego załącznika.
    Adres: /baza/attachments/<attachment_id>/delete/
    """

    attachment = get_object_or_404(EquipmentAttachment, pk=attachment_id)
    equipment_pk = attachment.equipment.pk

    if request.method == "POST":
        attachment.delete()
        return redirect("equipment:equipment_detail", pk=equipment_pk)

    return HttpResponseForbidden("Niedozwolone żądanie.")
